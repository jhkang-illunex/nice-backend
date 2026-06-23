"""외생충격 등비급수 — 매트릭스 전파 + 라운드별 수렴/발산 엑셀 생성.

대상: 현대모비스(1018116406) + 포스코(3018702315) 시드, 매입(upstream), depth 3,
      전체연도 합산 (실 company_edge 데이터). SCC 필터 없음(전체 depth-3 확장 그대로).

수록 내용
  1) 개요 — 시드/방향/연도/충격량/등비급수 공식 (SCC 통계 미기재)
  2) 전파행렬_R(희소) — R 의 비영 원소 747개를 (from,to,rate) triplet 로. R 은 320×320.
  3) 전파_1·2회 — v0(초기충격) → v1=R·v0 (1회) → v2=R²·v0 (2회) 를 노드별 벡터로.
                  맨 위에 Round-1 밀집 블록(시드 2 × 대상 19)을 두어 R·v 를 손으로 검산 가능하게.
  4) damping_1.0(발산) — 라운드별 등비급수 항/누적합 (잔여 안 줄어 발산)
  5) damping_0.85(수렴) — 라운드별 등비급수 항/누적합 (유한 수렴)

행렬 R 은 damping=1.0 기준. damping=α 이면 모든 원소가 ×α (rate = α·amt/Σ_out).
"""
from __future__ import annotations

from collections import defaultdict

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from nice_graph.shock import assemble as A

SEEDS = [("1018116406", None), ("3018702315", None)]
OUT = "docs/reports/shock/등비급수_전파행렬_매입전체_20260623.xlsx"
EPS = 1e-9
MAX_ROUNDS = 500  # 발산 케이스 표시 상한


def build(damping: float) -> A.PropagationInput:
    return A.assemble_propagation_input(
        SEEDS, depth=3, trade_year=None, within_subgraph=True,
        damping=damping, seed_shock=1.0, direction="upstream", normalize="source",
    )


def out_adj(pi: A.PropagationInput) -> dict[str, list[tuple[str, float]]]:
    adj: dict[str, list[tuple[str, float]]] = defaultdict(list)
    for e in pi.edges:
        adj[e["from_bizno"]].append((e["to_bizno"], e["rate"]))
    return adj


def step(adj, v: dict[str, float]) -> dict[str, float]:
    nx: dict[str, float] = defaultdict(float)
    for s, val in v.items():
        for d, r in adj.get(s, ()):
            nx[d] += val * r
    return nx


def round_by_round(pi: A.PropagationInput, max_rounds: int):
    """각 라운드 항 Σ(Rᵏ·init), 누적합, 항 max|값| 반환. (수렴 라운드 or None)."""
    adj = out_adj(pi)
    v0 = dict(pi.init_sub_graph)
    base = sum(v0.values())
    rows = [(0, base, base, max(abs(x) for x in v0.values()))]  # k=0 초기충격
    cum = base
    v = v0
    conv = None
    for k in range(1, max_rounds + 1):
        v = step(adj, v)
        term = sum(v.values())
        mx = max((abs(x) for x in v.values()), default=0.0)
        cum += term
        rows.append((k, term, cum, mx))
        if mx <= EPS:
            conv = k
            break
    return rows, conv


# ── 라벨 맵 ────────────────────────────────────────────────────────────────
def label_map(pi: A.PropagationInput) -> dict[str, str]:
    m = {}
    for n in pi.nodes:
        nm = (n.korentrnm or "").strip() or n.bizno
        m[n.node_id] = nm
    return m


# ── 스타일 헬퍼 ─────────────────────────────────────────────────────────────
TITLE = Font(bold=True, size=13)
HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="305496")
SEED_FILL = PatternFill("solid", fgColor="FCE4D6")
WARN = Font(bold=True, color="C00000")
OKF = Font(bold=True, color="2E7D32")


def style_head(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center")


def write_cycle_example(wb, pi, lab, adj) -> None:
    """닫힌 말단 고리(3사) 예제 시트 — 출구 없는 순환이 발산을 만드는 최소 사례.

    실데이터에서 현대종합기계·네오텍스·우림전열 3노드의 실제 거래비율을 추출.
    damping=1.0(충격 보존→발산) vs 0.85(매 홉 감쇠→수렴) 라운드 추적을 나란히 둔다.
    """
    NAMES = ["현대종합기계", "네오텍스", "우림전열"]
    ids = [nid for nid, nm in lab.items() if any(n in nm for n in NAMES)]
    ids = sorted(set(ids))
    # 고리 내부 엣지만 (damping=1.0 기준 rate). source 정규화라 행합=1.0.
    idset = set(ids)
    cyc_out = {s: [(d, r) for d, r in adj.get(s, ()) if d in idset] for s in ids}
    short = {nid: lab.get(nid, nid) for nid in ids}

    ws = wb.create_sheet("닫힌고리_예제(3사)")
    ws["A1"] = "닫힌 말단 고리 — 출구 없는 순환이 발산을 만드는 최소 실데이터 예제"
    ws["A1"].font = TITLE
    ws["A2"] = ("전체 거래망의 '말단(나가는 엣지 0)+순환' 성분. 이 고리는 sink로 빠져나갈 "
                "출구가 없어 충격이 영구히 갇힌다. ρ(R)=1 의 직접 원인.")
    ws["A3"] = ("개별 rate 는 모두 0<rate≤1 이지만, 각 노드 행합 Σ_out=1.0(=충격 100% 분배). "
                "rate<1 은 '감쇠'가 아니라 '분배'일 뿐 → 닫힌 고리에서 보존 → 발산.")

    # ① 엣지/행렬 (damping=1.0)
    r = 5
    ws.cell(row=r, column=1, value="① 고리 거래비율 R (damping=1.0, 실데이터)").font = Font(bold=True)
    r += 1
    hdr = ["from", "to", "rate(α=1.0)", "rate(α=0.85)"]
    for c, h in enumerate(hdr, 1):
        ws.cell(row=r, column=c, value=h)
    style_head(ws, r, len(hdr))
    r += 1
    for s in ids:
        for d, rate in sorted(cyc_out[s], key=lambda x: -x[1]):
            ws.cell(row=r, column=1, value=short[s])
            ws.cell(row=r, column=2, value=short[d])
            ws.cell(row=r, column=3, value=round(rate, 4))
            ws.cell(row=r, column=4, value=round(rate * 0.85, 4))
            r += 1
    # 행합 확인
    ws.cell(row=r, column=1, value="행합 Σ_out (각 노드)").font = Font(bold=True)
    r += 1
    for s in ids:
        tot = sum(rt for _, rt in cyc_out[s])
        ws.cell(row=r, column=1, value=short[s])
        ws.cell(row=r, column=2, value="Σ_out =")
        ws.cell(row=r, column=3, value=round(tot, 4))
        ws.cell(row=r, column=4, value=round(tot * 0.85, 4))
        r += 1
    r += 1

    # ②③ 라운드 추적 (충격 1.0 을 첫 노드에 투입)
    start = ids[0]  # 현대종합기계(분배 노드)일 가능성 높음; 라벨로 정렬 무관하게 동작

    def trace(alpha: float, max_k: int):
        v = {n: 0.0 for n in ids}
        v[start] = 1.0
        rows = [(0, dict(v), 1.0, 1.0)]
        cum = 1.0
        for k in range(1, max_k + 1):
            nv = {n: 0.0 for n in ids}
            for s in ids:
                for d, rate in cyc_out[s]:
                    nv[d] += v[s] * rate * alpha
            tot = sum(nv.values())
            cum += tot
            rows.append((k, dict(nv), tot, cum))
            v = nv
            if tot <= 1e-9:
                break
        return rows

    def write_trace(title_row, alpha, diverge, max_k):
        rr = title_row
        tag = "발산(고리합 1.0 유지)" if diverge else "수렴(고리합 0.85ᵏ로 감소)"
        ws.cell(row=rr, column=1,
                value=f"{'②' if diverge else '③'} 라운드 추적 — damping={alpha}  [{tag}]  (충격 1.0 → {short[start]})").font = Font(bold=True)
        rr += 1
        hdr = ["k"] + [short[n] for n in ids] + ["고리합", "누적Σ"]
        for c, h in enumerate(hdr, 1):
            ws.cell(row=rr, column=c, value=h)
        style_head(ws, rr, len(hdr))
        head = rr
        rr += 1
        rows = trace(alpha, max_k)
        for k, vec, tot, cum in rows:
            ws.cell(row=rr, column=1, value=k)
            for c, n in enumerate(ids, start=2):
                ws.cell(row=rr, column=c, value=round(vec[n], 6) or None)
            ws.cell(row=rr, column=2 + len(ids), value=round(tot, 6))
            ws.cell(row=rr, column=3 + len(ids), value=round(cum, 6))
            rr += 1
        last = rr - 1
        if diverge:
            ws.cell(row=rr, column=1,
                    value=f"→ 고리합이 매 라운드 1.0 그대로(보존). 누적Σ 무한 증가 → 발산.").font = WARN
        else:
            ws.cell(row=rr, column=1,
                    value=f"→ 고리합 0.85ᵏ로 0 수렴. 누적Σ={rows[-1][3]:.4f} (유한) → 수렴.").font = OKF
        # 차트: 고리합
        ch = LineChart()
        ch.title = f"damping={alpha} 고리합 추이"
        ch.height, ch.width = 6, 12
        data = Reference(ws, min_col=2 + len(ids), max_col=2 + len(ids), min_row=head, max_row=last)
        cats = Reference(ws, min_col=1, min_row=head + 1, max_row=last)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        ws.add_chart(ch, f"{get_column_letter(5 + len(ids))}{head}")
        return rr + 2

    nxt = write_trace(r, 1.0, diverge=True, max_k=12)
    write_trace(nxt, 0.85, diverge=False, max_k=80)

    for col, w in zip("ABCDE", (16, 16, 13, 13, 12)):
        ws.column_dimensions[col].width = w


def main() -> None:
    pi = build(1.0)              # 행렬/벡터 = damping 1.0 기준
    lab = label_map(pi)
    adj = out_adj(pi)
    v0 = dict(pi.init_sub_graph)
    v1 = step(adj, v0)
    v2 = step(adj, v1)

    def name(nid: str) -> str:
        return f"{lab.get(nid, nid)} [{nid}]"

    wb = Workbook()

    # ── 1) 개요 ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "개요"
    ws["A1"] = "외생충격 등비급수 — 매트릭스 전파 (매입·전체연도, 실데이터)"
    ws["A1"].font = TITLE
    meta = [
        ("시드", "현대모비스(주)(1018116406), (주)포스코(3018702315)"),
        ("방향", "매입 파급(upstream) — company_edge 역방향"),
        ("거래연도", "전체연도 합산(SUM)"),
        ("확장 깊이", "depth 3 (전체 확장 그대로, 별도 필터 없음)"),
        ("충격량(seed_shock)", "1.0 (시드별 균등)"),
        ("서브그래프 규모", f"노드 {len(pi.nodes)} · 엣지 {len(pi.edges)}"),
        ("전파행렬 R", f"{len(pi.nodes)}×{len(pi.nodes)} 행렬, 비영 원소 {len(pi.edges)}개 (damping=1.0 기준)"),
        ("등비급수", "total = Σ_k Rᵏ·init  (k=0 초기충격, k행 = Rᵏ·init 항)"),
        ("rate 정의", "rate = damping · amt / Σ_out(source)  → Σ_out = damping"),
        ("damping=α 효과", "행렬 R 의 모든 원소 ×α. α=1 이면 Σ_out=1 → 순환에서 충격 무손실 → 발산"),
    ]
    r = 3
    for k, v in meta:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="결론").font = WARN
    ws.cell(row=r, column=2,
            value="damping=1.0 → 발산(잔여항 안 줄고 누적합 무한 증가). "
                  "damping=0.85 → 유한 수렴. 수렴 보장을 위해 damping<1 필수.").font = WARN
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 86

    # ── 2) 전파행렬_R(희소) ──────────────────────────────────────────────
    ws = wb.create_sheet("전파행렬_R(희소)")
    ws["A1"] = f"전파행렬 R (희소표현) — {len(pi.nodes)}×{len(pi.nodes)}, 비영 {len(pi.edges)}개 · damping=1.0"
    ws["A1"].font = TITLE
    ws["A2"] = "R[from→to] = rate. v_{k+1}[to] = Σ_from R[from→to]·v_k[from]. damping=α 면 rate×α."
    hdr = ["from(소스)", "to(대상)", "rate", "from_id", "to_id"]
    for c, h in enumerate(hdr, 1):
        ws.cell(row=4, column=c, value=h)
    style_head(ws, 4, len(hdr))
    edges_sorted = sorted(pi.edges, key=lambda e: e["rate"], reverse=True)
    rr = 5
    for e in edges_sorted:
        ws.cell(row=rr, column=1, value=lab.get(e["from_bizno"], e["from_bizno"]))
        ws.cell(row=rr, column=2, value=lab.get(e["to_bizno"], e["to_bizno"]))
        ws.cell(row=rr, column=3, value=round(e["rate"], 8))
        ws.cell(row=rr, column=4, value=e["from_bizno"])
        ws.cell(row=rr, column=5, value=e["to_bizno"])
        rr += 1
    for col, w in zip("ABCDE", (30, 30, 12, 20, 20)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

    # ── 3) 전파_1·2회 (벡터 + Round1 밀집블록) ────────────────────────────
    ws = wb.create_sheet("전파_1·2회")
    ws["A1"] = "전파 이터레이터 — v0(초기충격) → v1=R·v0 (1회) → v2=R²·v0 (2회)"
    ws["A1"].font = TITLE

    # 3a) Round-1 밀집 블록: 행=시드(2), 열=round1 대상(19)
    r1_targets = sorted(v1, key=lambda n: v1[n], reverse=True)
    ws.cell(row=3, column=1,
            value="① Round-1 밀집행렬  R[시드→대상]  (행=시드, 열=v1 대상). v1 = 행별 (rate·v0) 합")
    ws.cell(row=3, column=1).font = Font(bold=True)
    # 헤더: 열 = 대상 노드명
    ws.cell(row=4, column=1, value="seed \\ target")
    for j, t in enumerate(r1_targets, start=2):
        ws.cell(row=4, column=j, value=lab.get(t, t))
        ws.cell(row=4, column=j).alignment = Alignment(textRotation=60, horizontal="center")
    style_head(ws, 4, 1 + len(r1_targets))
    seed_ids = list(v0)
    rowi = 5
    for s in seed_ids:
        ws.cell(row=rowi, column=1, value=f"{lab.get(s, s)} (v0={v0[s]:g})").fill = SEED_FILL
        srow = dict(adj.get(s, ()))
        for j, t in enumerate(r1_targets, start=2):
            val = srow.get(t, 0.0) * v0[s]
            if val:
                ws.cell(row=rowi, column=j, value=round(val, 6))
        rowi += 1
    # v1 합계행
    ws.cell(row=rowi, column=1, value="v1 = Σ (1회 전파)").font = Font(bold=True)
    for j, t in enumerate(r1_targets, start=2):
        ws.cell(row=rowi, column=j, value=round(v1[t], 6)).font = Font(bold=True)
    rowi += 2

    # 3b) 노드별 벡터 표 (v0/v1/v2)
    ws.cell(row=rowi, column=1,
            value="② 노드별 전파벡터 (v1=1회 활성 19, v2=2회 활성 313 中 v2 상위 표시)")
    ws.cell(row=rowi, column=1).font = Font(bold=True)
    rowi += 1
    hdr2 = ["노드", "v0 (초기충격)", "v1 = R·v0 (1회)", "v2 = R²·v0 (2회)", "node_id"]
    for c, h in enumerate(hdr2, 1):
        ws.cell(row=rowi, column=c, value=h)
    style_head(ws, rowi, len(hdr2))
    head_row = rowi
    rowi += 1
    union = set(v0) | set(v1) | set(v2)
    # 정렬: v1 우선, 그다음 v2
    order = sorted(union, key=lambda n: (v1.get(n, 0.0), v2.get(n, 0.0)), reverse=True)
    for n in order:
        ws.cell(row=rowi, column=1, value=lab.get(n, n))
        ws.cell(row=rowi, column=2, value=round(v0.get(n, 0.0), 6) or None)
        ws.cell(row=rowi, column=3, value=round(v1.get(n, 0.0), 6) or None)
        ws.cell(row=rowi, column=4, value=round(v2.get(n, 0.0), 6) or None)
        ws.cell(row=rowi, column=5, value=n)
        if n in v0:
            for c in range(1, 6):
                ws.cell(row=rowi, column=c).fill = SEED_FILL
        rowi += 1
    # 합계
    ws.cell(row=rowi, column=1, value="합계 Σ").font = Font(bold=True)
    ws.cell(row=rowi, column=2, value=round(sum(v0.values()), 6)).font = Font(bold=True)
    ws.cell(row=rowi, column=3, value=round(sum(v1.values()), 6)).font = Font(bold=True)
    ws.cell(row=rowi, column=4, value=round(sum(v2.values()), 6)).font = Font(bold=True)
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
    for col, w in zip("ABCDE", (34, 14, 16, 16, 18)):
        ws.column_dimensions[col].width = w

    # ── 4·5) 라운드별 수렴/발산 ──────────────────────────────────────────
    def write_rounds(title: str, damping: float, diverge: bool):
        rows, conv = round_by_round(build(damping), MAX_ROUNDS)
        ws = wb.create_sheet(title)
        tag = "발산" if diverge else "수렴"
        ws["A1"] = f"damping={damping} — 라운드별 등비급수 (매입·전체연도) [{tag}]"
        ws["A1"].font = TITLE
        hdr = ["라운드 k", "이번 항 Σ(Rᵏ·init)", "누적 합계 Σ_0..k", "항 max|값|"]
        for c, h in enumerate(hdr, 1):
            ws.cell(row=3, column=c, value=h)
        style_head(ws, 3, len(hdr))
        for i, (k, term, cum, mx) in enumerate(rows, start=4):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=round(term, 6))
            ws.cell(row=i, column=3, value=round(cum, 6))
            ws.cell(row=i, column=4, value=round(mx, 8))
        last = 4 + len(rows) - 1
        note_row = last + 2
        if conv:
            ws.cell(row=note_row, column=1,
                    value=f"→ 항이 0으로 수렴 → {conv}회 종료. 누적합 수렴값 Σ={rows[-1][2]:.4f} (유한)").font = OKF
        else:
            ws.cell(row=note_row, column=1,
                    value=f"→ 항이 안 줄어 {MAX_ROUNDS}라운드까지 발산. "
                          f"누적합 Σ={rows[-1][2]:.4f}이며 계속 증가(무한)").font = WARN
        # 차트: 누적합 + 항max
        ch = LineChart()
        ch.title = f"damping={damping} 누적합 / 잔여항"
        ch.height, ch.width = 8, 16
        data = Reference(ws, min_col=3, max_col=4, min_row=3, max_row=last)
        cats = Reference(ws, min_col=1, min_row=4, max_row=last)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        ws.add_chart(ch, "F3")
        for col, w in zip("ABCD", (10, 20, 18, 14)):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A4"

    write_rounds("damping_1.0(발산)", 1.0, diverge=True)
    write_rounds("damping_0.85(수렴)", 0.85, diverge=False)

    write_cycle_example(wb, pi, lab, adj)

    wb.save(OUT)
    print(f"saved: {OUT}")
    print(f"  노드 {len(pi.nodes)} · 엣지 {len(pi.edges)} · v1활성 {len(v1)} · v2활성 {len(v2)}")
    print(f"  Σv0={sum(v0.values()):.4f}  Σv1={sum(v1.values()):.4f}  Σv2={sum(v2.values()):.4f}")


if __name__ == "__main__":
    main()
