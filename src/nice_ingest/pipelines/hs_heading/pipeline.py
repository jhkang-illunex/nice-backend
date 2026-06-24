"""관세청 'HS부호 단위별 품목명' xlsx → rag.hs_heading 적재.

원천: 공공데이터포털 데이터셋 15130660 (관세청_HS부호 단위별 품목명).
시트 구성: HS2단위 / HS4단위 / HS6단위(5단위포함) / HS8단위(7, 9단위포함) /
HS10단위 — 각 시트 [부호, 한글품목명, 영문품목명].

수집 자동화 (인터넷 가능 구간에서 1회):
  curl -L "https://www.data.go.kr/cmm/cmm/fileDownload.do?atchFileId=FILE_000000003584116&fileDetailSn=1" \\
       -o 관세청_HS부호단위별품목명.xlsx
  (atchFileId 는 고시 개정 시 바뀔 수 있음 — 데이터셋 페이지에서 재확인)

적재 규칙
  - prefix 길이 = level (2~10). 10단위는 rag.hsk.name_ko 와 중복이지만
    정합성 대조용으로 함께 보관.
  - 멱등: INSERT ... ON CONFLICT (hs_prefix) DO UPDATE.
  - 검증: 레벨별 건수 + rag.hsk 와의 4단위 커버리지 출력.
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path
from typing import Any

log = logging.getLogger(__name__)

_UPSERT = """
    INSERT INTO rag.hs_heading (hs_prefix, level, name_ko, name_en, updated_at)
    VALUES (%(hs_prefix)s, %(level)s, %(name_ko)s, %(name_en)s, now())
    ON CONFLICT (hs_prefix) DO UPDATE
    SET level = EXCLUDED.level,
        name_ko = EXCLUDED.name_ko,
        name_en = EXCLUDED.name_en,
        updated_at = now()
"""

_VERIFY = """
    SELECT g.level, count(*) AS n
    FROM rag.hs_heading g GROUP BY g.level ORDER BY g.level
"""

_COVERAGE = """
    SELECT count(*) AS total,
           count(*) FILTER (WHERE EXISTS (
               SELECT 1 FROM rag.hs_heading g WHERE g.hs_prefix = LEFT(h.hs_code, 4)
           )) AS hs4_covered
    FROM rag.hsk h
"""


def iter_rows(path: Path) -> list[dict[str, Any]]:
    # openpyxl 은 ingest extras (.[ingest]) — hscode 파이프라인과 동일 정책
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    out: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            prefix = str(r[0]).strip()
            if not prefix.isdigit() or not (2 <= len(prefix) <= 10):
                continue
            name_ko = str(r[1]).strip() if r[1] is not None else None
            name_en = str(r[2]).strip() if len(r) > 2 and r[2] is not None else None
            if not name_ko and not name_en:
                continue
            out.append(
                {
                    "hs_prefix": prefix,
                    "level": len(prefix),
                    "name_ko": name_ko,
                    "name_en": name_en,
                }
            )
    return out


def ingest(path: Path, *, dry_run: bool = False) -> int:
    rows = iter_rows(path)
    levels: dict[int, int] = {}
    for r in rows:
        levels[r["level"]] = levels.get(r["level"], 0) + 1
    print(f"파싱: {len(rows)}행, 레벨별 {dict(sorted(levels.items()))}")
    if dry_run:
        return 0

    import psycopg

    from nice_common.config import get_settings

    s = get_settings()
    dsn = (
        f"host={s.postgres_host} port={s.postgres_port} user={s.postgres_user} "
        f"password={s.postgres_password} dbname={s.postgres_db}"
    )
    with psycopg.connect(dsn) as conn:
        with conn.cursor() as cur:
            cur.executemany(_UPSERT, rows)
        conn.commit()
        with conn.cursor() as cur:
            cur.execute(_VERIFY)
            for level, n in cur.fetchall():
                print(f"  level {level:2d}: {n}건")
            cur.execute(_COVERAGE)
            total, covered = cur.fetchone()
            print(f"rag.hsk 4단위 커버리지: {covered}/{total}")
    return 0


def add_args(p: argparse.ArgumentParser) -> None:
    p.add_argument("--file", required=True, help="단위별 품목명 xlsx 경로")
    p.add_argument("--dry-run", action="store_true", help="파싱 통계만 출력 (DB 미접속)")


def run(args: argparse.Namespace) -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    try:
        return ingest(Path(args.file), dry_run=args.dry_run)
    except Exception as exc:  # noqa: BLE001
        print(f"hs_heading ingest failed: {exc.__class__.__name__}: {exc}", file=sys.stderr)
        return 1
