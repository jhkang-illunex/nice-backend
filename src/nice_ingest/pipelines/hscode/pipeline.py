"""HSCode 적재 파이프라인 — 관세청 HS부호 xlsx → PostgreSQL ``hsk``.

흐름
  1. openpyxl 의 read_only 스트리밍으로 12k row 를 메모리 폭주 없이 순회
  2. 헤더 → DB 컬럼 매핑 + 행 단위 변환(빈문자→None, 날짜→date, hs_code 10자리 zero-pad)
  3. ``--active-only`` 시 ``valid_to >= today`` 필터
  4. psycopg3 executemany 로 ``INSERT ... ON CONFLICT (hs_code) DO UPDATE`` (멱등)
  5. ``--dry-run`` 시 DB 미접속, 파싱/통계만

임베딩(``embedding`` 컬럼)은 본 파이프라인에서 채우지 않는다 — 별 단계에서
``nice_rag.search.hsk_embed`` 가 배치 임베딩 후 UPDATE.
"""

from __future__ import annotations

import argparse
import datetime as dt
import logging
import sys
from collections.abc import Iterable, Iterator
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

log = logging.getLogger(__name__)


# Excel 헤더(한글) → DB 컬럼명 매핑. 21번째 빈 컬럼은 무시.
HEADER_TO_COLUMN: dict[str, str] = {
    "HS부호": "hs_code",
    "적용시작일자": "valid_from",
    "적용종료일자": "valid_to",
    "한글품목명": "name_ko",
    "영문품목명": "name_en",
    "HS부호내용": "hs_content",
    "한국표준무역분류명": "standard_trade_name",
    "수량단위최대단가": "qty_unit_max_price",
    "중량단위최대단가": "weight_unit_max_price",
    "수량단위코드": "qty_unit_code",
    "중량단위코드": "weight_unit_code",
    "수출성질코드": "export_nature_code",
    "수입성질코드": "import_nature_code",
    "품목규격명": "item_spec_name",
    "필수규격명": "required_spec_name",
    "참고규격명": "ref_spec_name",
    "규격설명": "spec_description",
    "규격사항내용": "spec_content",
    "성질통합분류코드": "nature_integrated_code",
    "성질통합분류코드명": "nature_integrated_name",
}

# upsert 대상 컬럼 순서 (SQL placeholder 매칭에 사용)
COLUMNS: list[str] = list(HEADER_TO_COLUMN.values())

UPSERT_SQL = """
INSERT INTO rag.hsk (
    hs_code, valid_from, valid_to,
    name_ko, name_en, hs_content, standard_trade_name,
    qty_unit_max_price, weight_unit_max_price,
    qty_unit_code, weight_unit_code,
    export_nature_code, import_nature_code,
    item_spec_name, required_spec_name, ref_spec_name,
    spec_description, spec_content,
    nature_integrated_code, nature_integrated_name
) VALUES (
    %(hs_code)s, %(valid_from)s, %(valid_to)s,
    %(name_ko)s, %(name_en)s, %(hs_content)s, %(standard_trade_name)s,
    %(qty_unit_max_price)s, %(weight_unit_max_price)s,
    %(qty_unit_code)s, %(weight_unit_code)s,
    %(export_nature_code)s, %(import_nature_code)s,
    %(item_spec_name)s, %(required_spec_name)s, %(ref_spec_name)s,
    %(spec_description)s, %(spec_content)s,
    %(nature_integrated_code)s, %(nature_integrated_name)s
)
ON CONFLICT (hs_code) DO UPDATE SET
    valid_from              = EXCLUDED.valid_from,
    valid_to                = EXCLUDED.valid_to,
    name_ko                 = EXCLUDED.name_ko,
    name_en                 = EXCLUDED.name_en,
    hs_content              = EXCLUDED.hs_content,
    standard_trade_name     = EXCLUDED.standard_trade_name,
    qty_unit_max_price      = EXCLUDED.qty_unit_max_price,
    weight_unit_max_price   = EXCLUDED.weight_unit_max_price,
    qty_unit_code           = EXCLUDED.qty_unit_code,
    weight_unit_code        = EXCLUDED.weight_unit_code,
    export_nature_code      = EXCLUDED.export_nature_code,
    import_nature_code      = EXCLUDED.import_nature_code,
    item_spec_name          = EXCLUDED.item_spec_name,
    required_spec_name      = EXCLUDED.required_spec_name,
    ref_spec_name           = EXCLUDED.ref_spec_name,
    spec_description        = EXCLUDED.spec_description,
    spec_content            = EXCLUDED.spec_content,
    nature_integrated_code  = EXCLUDED.nature_integrated_code,
    nature_integrated_name  = EXCLUDED.nature_integrated_name;
"""


@dataclass
class IngestReport:
    total_rows: int = 0
    parsed_rows: int = 0
    skipped: dict[str, int] = field(default_factory=dict)
    active_rows: int | None = None  # --active-only 모드일 때만
    upserted: int = 0
    dry_run: bool = False

    def bump_skip(self, reason: str) -> None:
        self.skipped[reason] = self.skipped.get(reason, 0) + 1

    def summary(self) -> str:
        lines = [
            f"total rows         : {self.total_rows}",
            f"parsed             : {self.parsed_rows}",
            f"skipped            : {sum(self.skipped.values())} {self.skipped or ''}",
        ]
        if self.active_rows is not None:
            lines.append(f"active (after filt): {self.active_rows}")
        lines.append(f"upserted           : {self.upserted}{' (dry-run)' if self.dry_run else ''}")
        return "\n".join(lines)


# ─── 순수 변환 함수 (테스트 가능, 외부 의존 없음) ────────────────────────────

def _to_text(v: Any) -> str | None:
    """공백 trim + 빈문자열을 NULL 로 정규화."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_date(v: Any) -> dt.date | None:
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    if not s:
        return None
    # 관세청 포맷이 ISO 면 fromisoformat 으로 충분 — 다른 포맷이면 호출 측이 확장
    try:
        return dt.date.fromisoformat(s[:10])
    except ValueError:
        return None


def _to_numeric(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _normalize_hs_code(v: Any) -> str | None:
    """HS부호 → 10자리 문자열 정규화. 비숫자 / 길이 초과는 None.

    관세청 파일의 HS부호 셀은 전량 텍스트라 선행 0 이 보존되고, 10자리 미만은
    후행 0 이 생략된 형태다 — 따라서 텍스트 셀은 오른쪽 0 패딩만 한다.
    예: '01059910' → 0105991000, '271020974' → 2710209740, '0507901' → 0507901000.
    왼쪽 패딩(zfill)은 '0001059910'(존재하지 않는 00류, 987행) 과
    '0271020974'(02류로 위장한 유령 코드, 9자리 셀) 오염의 원인이었다.

    숫자형 셀은 선행 0 이 소실된 경우라 홀수 길이면 1개 복원 후 우측 패딩
    (현 고시본에는 숫자형 셀이 없지만 포맷 변경 대비 방어).
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        s = str(int(v))
        if len(s) % 2 == 1:
            s = "0" + s
    else:
        s = str(v).strip()
    if not s.isdigit():
        return None
    if len(s) > 10:
        return None
    return s.ljust(10, "0")


def transform_row(raw: dict[str, Any]) -> dict[str, Any] | None:
    """헤더-매핑된 1행 dict → 적재용 dict. 적재 불가하면 None."""
    hs_code = _normalize_hs_code(raw.get("hs_code"))
    if hs_code is None:
        return None
    valid_from = _to_date(raw.get("valid_from"))
    valid_to = _to_date(raw.get("valid_to"))
    if valid_from is None or valid_to is None:
        return None

    return {
        "hs_code": hs_code,
        "valid_from": valid_from,
        "valid_to": valid_to,
        "name_ko": _to_text(raw.get("name_ko")),
        "name_en": _to_text(raw.get("name_en")),
        "hs_content": _to_text(raw.get("hs_content")),
        "standard_trade_name": _to_text(raw.get("standard_trade_name")),
        "qty_unit_max_price": _to_numeric(raw.get("qty_unit_max_price")),
        "weight_unit_max_price": _to_numeric(raw.get("weight_unit_max_price")),
        "qty_unit_code": _to_text(raw.get("qty_unit_code")),
        "weight_unit_code": _to_text(raw.get("weight_unit_code")),
        "export_nature_code": _to_text(raw.get("export_nature_code")),
        "import_nature_code": _to_text(raw.get("import_nature_code")),
        "item_spec_name": _to_text(raw.get("item_spec_name")),
        "required_spec_name": _to_text(raw.get("required_spec_name")),
        "ref_spec_name": _to_text(raw.get("ref_spec_name")),
        "spec_description": _to_text(raw.get("spec_description")),
        "spec_content": _to_text(raw.get("spec_content")),
        "nature_integrated_code": _to_text(raw.get("nature_integrated_code")),
        "nature_integrated_name": _to_text(raw.get("nature_integrated_name")),
    }


# ─── Excel 스트리밍 ──────────────────────────────────────────────────────────

def iter_excel_rows(path: Path) -> Iterator[dict[str, Any]]:
    """xlsx 의 첫 시트를 헤더 매핑된 dict 로 yield (스트리밍)."""
    # openpyxl 은 ingest 컨테이너에만 설치(.[ingest] extras). 호스트 dev 에서
    # ``pip install -e .[ingest]`` 가 안 된 상태면 ImportError 가 명확히 노출.
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return
        # 인덱스 → DB 컬럼 (모르는 헤더는 None)
        idx_to_col: list[str | None] = [
            HEADER_TO_COLUMN.get(str(h).strip() if h is not None else "") for h in header_row
        ]
        for raw in rows:
            yield {
                col: raw[i]
                for i, col in enumerate(idx_to_col)
                if col is not None and i < len(raw)
            }
    finally:
        wb.close()


# ─── 적재 ────────────────────────────────────────────────────────────────────

def _filter_active(rows: Iterable[dict[str, Any]], today: dt.date) -> Iterator[dict[str, Any]]:
    for r in rows:
        if r["valid_to"] >= today:
            yield r


def _executemany_upsert(rows: list[dict[str, Any]]) -> int:
    """psycopg3 로 일괄 upsert. 반환값 = 적재 시도 row 수."""
    import psycopg

    from nice_common.config import get_settings

    dsn = get_settings().postgres_dsn
    # SQLAlchemy DSN(`postgresql+psycopg://...`) 을 psycopg 가 이해하는 형태로 정규화
    if dsn.startswith("postgresql+psycopg://"):
        dsn = "postgresql://" + dsn[len("postgresql+psycopg://"):]

    with psycopg.connect(dsn) as conn:
        with conn.cursor() as cur:
            cur.executemany(UPSERT_SQL, rows)
        conn.commit()
    return len(rows)


def ingest(
    *,
    path: Path,
    active_only: bool = False,
    dry_run: bool = False,
    today: dt.date | None = None,
) -> IngestReport:
    """공개 API — CLI / 테스트 양쪽에서 호출."""
    report = IngestReport(dry_run=dry_run)
    today = today or dt.date.today()

    accepted: list[dict[str, Any]] = []
    for raw in iter_excel_rows(path):
        report.total_rows += 1
        row = transform_row(raw)
        if row is None:
            report.bump_skip("invalid_hs_code_or_dates")
            continue
        report.parsed_rows += 1
        accepted.append(row)

    if active_only:
        before = len(accepted)
        accepted = list(_filter_active(accepted, today))
        report.active_rows = len(accepted)
        if before != len(accepted):
            report.skipped["expired"] = before - len(accepted)

    if dry_run:
        report.upserted = len(accepted)
        return report

    if not accepted:
        return report

    report.upserted = _executemany_upsert(accepted)
    return report


# ─── CLI 등록 ────────────────────────────────────────────────────────────────

def add_args(p: argparse.ArgumentParser) -> None:
    p.add_argument(
        "--file",
        required=True,
        help="관세청 HS부호 Excel 경로 (예: /work/관세청_HS부호_20260101.xlsx)",
    )
    p.add_argument(
        "--active-only",
        action="store_true",
        help="valid_to >= today 인 row 만 적재 (기본 false — 전체)",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="DB 적재 없이 파싱/검증만",
    )


def run(args: argparse.Namespace) -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    path = Path(args.file)
    if not path.exists():
        print(f"file not found: {path}", file=sys.stderr)
        return 2
    report = ingest(path=path, active_only=args.active_only, dry_run=args.dry_run)
    print(report.summary())
    return 0
