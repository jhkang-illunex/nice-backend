"""KSIC 적재 파이프라인 — 한국표준산업분류(제11차) xlsx → PostgreSQL ``rag.ksic``.

원천 파일
  통계청 제11차 개정 분류체계 표 (docs/data/통계청_한국표준산업분류_제11차_분류체계.xlsx).
  wide 포맷 1시트: (대분류 코드·항목명, 중분류 …, 소분류 …, 세분류 …, 세세분류 …)
  10컬럼, 상위 계층 셀은 첫 등장 행에만 값이 있고 이후 빈 셀(forward-fill 구조).
  파싱 결과는 공식 항목 수(대21·중77·소234·세501·세세1,205)와 대조 검증한다.

적재 범위는 대분류(level=1)·중분류(level=2)의 **98 row 뿐**이다 — 소분류
이하는 row 로 넣지 않는 대신 항목명을 ``children_text`` 로 결합해 상위
row 의 검색 텍스트에 흡수시킨다 (예: '반도체' 질의가 중분류 26 에 걸리는
경로는 소분류 261 '반도체 제조업' 명칭이다).

임베딩(``embedding`` 컬럼)은 본 파이프라인에서 채우지 않는다 — 별 단계에서
``nice_ingest run ksic_embed`` 가 배치 임베딩 후 UPDATE. 재적재로
``search_text`` 가 바뀐 경우 ``ksic_embed --rebuild`` 로 재임베딩할 것.
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from collections.abc import Iterable, Iterator
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

log = logging.getLogger(__name__)

# 공식 제11차 항목 수 — 파싱 검증 기준 (통계청 고시 제2024-203호)
EXPECTED_COUNTS = {1: 21, 2: 77, 3: 234, 4: 501, 5: 1205}

UPSERT_SQL = """
INSERT INTO rag.ksic (code, level, parent_code, name_ko, division_range, children_text, search_text)
VALUES (%(code)s, %(level)s, %(parent_code)s, %(name_ko)s, %(division_range)s,
        %(children_text)s, %(search_text)s)
ON CONFLICT (code) DO UPDATE SET
    level           = EXCLUDED.level,
    parent_code     = EXCLUDED.parent_code,
    name_ko         = EXCLUDED.name_ko,
    division_range  = EXCLUDED.division_range,
    children_text   = EXCLUDED.children_text,
    search_text     = EXCLUDED.search_text;
"""


@dataclass
class IngestReport:
    level_counts: dict[int, int] = field(default_factory=dict)
    count_mismatch: dict[int, tuple[int, int]] = field(default_factory=dict)
    upserted: int = 0
    dry_run: bool = False

    def summary(self) -> str:
        label = {1: "대분류", 2: "중분류", 3: "소분류", 4: "세분류", 5: "세세분류"}
        lines = [
            f"{label[lv]:<4} : {self.level_counts.get(lv, 0)}"
            for lv in sorted(label)
        ]
        if self.count_mismatch:
            lines.append(f"MISMATCH vs 공식 항목 수: {self.count_mismatch}")
        lines.append(f"upserted : {self.upserted}{' (dry-run)' if self.dry_run else ''}")
        return "\n".join(lines)


# ─── 순수 변환 함수 (테스트 가능, 외부 의존 없음) ────────────────────────────

_SECTION_RE = re.compile(r"^(?P<name>.+?)\s*\((?P<range>\d{2}(?:~\d{2})?)\)$")


def split_section_name(raw: str) -> tuple[str, str | None]:
    """대분류 항목명 '농업, 임업 및 어업(01~03)' → ('농업, 임업 및 어업', '01~03')."""
    m = _SECTION_RE.match(raw.strip())
    if not m:
        return raw.strip(), None
    return m.group("name"), m.group("range")


def _dedup_keep_order(names: Iterable[str]) -> list[str]:
    """세분류·세세분류가 동명인 경우가 흔함('곡물 및 기타 식량작물 재배업') — 중복 제거."""
    seen: set[str] = set()
    out: list[str] = []
    for n in names:
        if n not in seen:
            seen.add(n)
            out.append(n)
    return out


def build_search_text(name_ko: str, parent_name: str | None, children: list[str]) -> str:
    """rag.hsk 의 search_text 와 같은 ' | ' 구분 포맷 — 임베딩·ts 색인의 단일 진실."""
    return f"{name_ko} | {parent_name or ''} | {' '.join(_dedup_keep_order(children))}"


def parse_rows(rows: Iterable[tuple[Any, ...]]) -> tuple[list[dict[str, Any]], dict[int, int]]:
    """wide 포맷 행 스트림 → 적재 레코드 리스트 + 계층별 항목 수.

    '코드' 헤더 행 이전은 건너뛴다. 상위 계층 셀은 forward-fill 로 추적.
    Returns (records, level_counts) — records 는 대분류가 중분류보다 먼저
    (parent_code FK 만족을 위해 삽입 순서 보장).
    """
    counts = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    sections: list[dict[str, Any]] = []
    divisions: list[dict[str, Any]] = []
    sec_children: dict[str, list[str]] = {}
    div_children: dict[str, list[str]] = {}
    cur_sec: str | None = None
    cur_div: str | None = None

    in_body = False
    for r in rows:
        cells = (list(r) + [None] * 10)[:10]
        a, an, b, bn, c, cn, d, dn, e, en = (
            str(v).strip() if v is not None else None for v in cells
        )
        if not in_body:
            in_body = a == "코드"
            continue

        if a is not None and an is not None:
            counts[1] += 1
            name, div_range = split_section_name(an)
            cur_sec = a
            sec_children[a] = []
            sections.append(
                {
                    "code": a,
                    "level": 1,
                    "parent_code": None,
                    "name_ko": name,
                    "division_range": div_range,
                }
            )
        if b is not None and bn is not None:
            counts[2] += 1
            if cur_sec is None:
                raise ValueError(f"중분류 {b}({bn}) 가 대분류보다 먼저 등장")
            cur_div = b
            div_children[b] = []
            sec_children[cur_sec].append(bn)
            divisions.append(
                {
                    "code": b,
                    "level": 2,
                    "parent_code": cur_sec,
                    "name_ko": bn,
                    "division_range": None,
                }
            )
        # 소분류 이하는 row 로 적재하지 않고 상위의 children_text 재료로만 사용.
        # 대분류는 중분류명+소분류명까지만 흡수 — 세(세세)분류까지 넣으면
        # 제조업(C) 이 700여 항목명으로 부풀어 임베딩 입력이 희석/초과된다.
        for lv, name in ((3, cn), (4, dn), (5, en)):
            if name is None:
                continue
            counts[lv] += 1
            if cur_sec is not None and lv == 3:
                sec_children[cur_sec].append(name)
            if cur_div is not None:
                div_children[cur_div].append(name)

    sec_name = {s["code"]: s["name_ko"] for s in sections}
    for s in sections:
        s["children_text"] = " ".join(_dedup_keep_order(sec_children[s["code"]]))
        s["search_text"] = build_search_text(s["name_ko"], None, sec_children[s["code"]])
    for dv in divisions:
        dv["children_text"] = " ".join(_dedup_keep_order(div_children[dv["code"]]))
        dv["search_text"] = build_search_text(
            dv["name_ko"], sec_name.get(dv["parent_code"]), div_children[dv["code"]]
        )
    return sections + divisions, counts


# ─── Excel 스트리밍 ──────────────────────────────────────────────────────────

def iter_excel_rows(path: Path) -> Iterator[tuple[Any, ...]]:
    """xlsx 첫 시트를 튜플 스트림으로 yield (hscode 파이프라인과 동일 정책)."""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        yield from ws.iter_rows(values_only=True)
    finally:
        wb.close()


# ─── 적재 ────────────────────────────────────────────────────────────────────

def _executemany_upsert(rows: list[dict[str, Any]]) -> int:
    import psycopg

    from nice_common.config import get_settings

    dsn = get_settings().postgres_dsn
    if dsn.startswith("postgresql+psycopg://"):
        dsn = "postgresql://" + dsn[len("postgresql+psycopg://"):]

    with psycopg.connect(dsn) as conn:
        with conn.cursor() as cur:
            cur.executemany(UPSERT_SQL, rows)
        conn.commit()
    return len(rows)


def ingest(*, path: Path, dry_run: bool = False) -> IngestReport:
    """공개 API — CLI / 테스트 양쪽에서 호출."""
    records, counts = parse_rows(iter_excel_rows(path))
    report = IngestReport(dry_run=dry_run, level_counts=counts)
    for lv, expected in EXPECTED_COUNTS.items():
        if counts.get(lv) != expected:
            report.count_mismatch[lv] = (counts.get(lv, 0), expected)
    if report.count_mismatch:
        log.warning(
            "파싱 항목 수가 공식 제11차 항목 수와 다름 (파일 포맷 변경?): %s",
            report.count_mismatch,
        )

    if dry_run:
        report.upserted = len(records)
        return report
    report.upserted = _executemany_upsert(records)
    return report


# ─── CLI 등록 ────────────────────────────────────────────────────────────────

def add_args(p: argparse.ArgumentParser) -> None:
    p.add_argument(
        "--file",
        required=True,
        help="제11차 분류체계 Excel 경로 (예: docs/data/통계청_한국표준산업분류_제11차_분류체계.xlsx)",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="DB 적재 없이 파싱/검증만",
    )


def run(args: argparse.Namespace) -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    path = Path(args.file)
    if not path.exists():
        print(f"file not found: {path}", file=sys.stderr)
        return 2
    report = ingest(path=path, dry_run=args.dry_run)
    print(report.summary())
    return 0
